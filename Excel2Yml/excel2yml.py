import openpyxl
from openpyxl.utils import get_column_letter
import ruamel.yaml
import sys
import os

strstr = r"""
- service: kerberos
  componentList:
    - component: master
      hostList: ["xinode2.local"]
    - component: slave
      hostList: ["xinode3.local"]
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
  configList:
    global:
      local_realm: "BDMS.COM"
    krb:
      domain_realm_maps: "{\"BDMS.COM\":[\"bdms.com\",\".bdms.com\"]}"

- service: ldap
  componentList:
    - component: server
      hostList: ["node1.local", "node2.local"]
    - component: client
      hostList: ["*"]
  configList:
    slapd-init.ldif:
      ldap_domain: "bdms.com"

# MySQL当前仅支持Centos7和RedHat7系统，且只能用于POC项目
- service: mysql
  componentList:
    - component: server
      hostList: ["node1.local"]

- service: easy_ranger
  componentList:
    - component: admin
      hostList: ["xinode2.local","xinode3.local"]
  configList:
    env:
      "JAVA_MEM_OPTS": "-XX:MetaspaceSize=100m -XX:MaxMetaspaceSize=200m -Xmx3g -Xms3g"

- service: zookeeper
  componentList:
    - component: server
      hostList: ["node{3-5}.local"]
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
  configList:
    env:
      JAVA_OPTS: "-server -Xmx3g -Xms3g -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -verbose:gc -XX:+PrintGCDetails -XX:+PrintGCTimeStamps -XX:+PrintGCDateStamps -XX:+PrintTenuringDistribution -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=3 -XX:GCLogFileSize=512M"  

- service: kafka
  componentList:
    - component: manager
      hostList: ["node4.local"]
    - component: broker
      # 必须安装≥2个节点，建议3个
      hostList: ["node{4-5}.local"]
  configList:
    env:
      heap_args: "-Xmx3g -Xms3g"
    manager:
      manager_heap_args: "-J-Xms1g -J-Xmx1g"  

- service: elasticsearch
  componentList:
    - component: master
      hostList: ["node{3-5}.local"]
    - component: data
      hostList: ["node{3-5}.local"]
  configList:
    data_jvm:
      jvm_heap_size: "4g"
    master_jvm:
      jvm_heap_size: "1g"

- service: neo4j
  componentList:
    - component: server
      hostList: ["xinode2.local", "xinode3.local", "xinode4.local"]

# -----------------Nginx和Ningx_HA二选一，正式环境必须用Ningx_HA-----------------
- service: nginx
  componentList:
    - component: server
      hostList: ["node2.local"]

- service: nginx_ha
  componentList:
    - component: server
      hostList: ["node3.local","node4.local"]
  configList:
    keepalived_common:
      default_virtual_ip_address: "<vip地址>"

- service: hdfs
  componentList:
    # zkfc 必须和 namenode 在相同机器
    - component: zkfc
      hostList: ["node1.local", "node2.local"]
    - component: namenode
      hostList: ["node1.local", "node2.local"]
    - component: journalnode
      hostList: ["node{1-3}.local"]
    - component: datanode
      hostList: ["node[1-5].local"]
    - component: client
      # 所有机器都安装client
      hostList: ["*"]

- service: yarn
  componentList:
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
    - component: nodemanager
      hostList: ["node[1-4].local"]
    - component: resourcemanager
      hostList: ["node3.local", "node4.local"]
    - component: historyserver
      hostList: ["node2.local"]
  configList:
    mapred_env:
      HADOOP_JOB_HISTORYSERVER_HEAPSIZE: 900
    yarn_env:
      "YARN_RESOURCEMANAGER_HEAPSIZE": "8192"
      "YARN_NODEMANAGER_HEAPSIZE": "4096"

- service: knox
  componentList:
    - component: server
      hostList: ["node3.local","node4.local"]
  configList:
    global:
      # nginx单点模式通过nginx ip 地址访问，不通过主机名访问
      domain: "<nginx对外域名或ip>:8889"
      
- service: hive
  componentList:
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
    - component: hiveserver
      hostList: ["node{1-3}.local"]
    - component: metastore
      hostList: ["node{4-5}.local"]
  configList:
    metastore:
      "hive_metastore_jvm_opts": "-Xmx12g -Xms12g -XX:PermSize=512m -XX:+UseConcMarkSweepGC -verbose:gc -XX:+PrintGCDetails -XX:+PrintGCDateStamps -XX:+PrintTenuringDistribution -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=10 -XX:GCLogFileSize=64M -XX:+HeapDumpOnOutOfMemoryError"  
    hiveserver:
      "hive_hiveserver_jvm_opts": "-Xmx18g -Xms18g -Xmn6000m -XX:MaxNewSize=6000m -XX:PermSize=2G -XX:+UseConcMarkSweepGC -verbose:gc -XX:+PrintGCDetails -XX:+PrintGCDateStamps -XX:+PrintTenuringDistribution -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=10 -XX:GCLogFileSize=64M"

- service: impala
  componentList:
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
    - component: catalogd
      hostList: ["node5.local"]
    - component: statestored
      hostList: ["node3.local"]
    - component: impalad
      hostList: ["node4.local"]

- service: spark2
  componentList:
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
    - component: jobhistoryserver
      hostList: ["node4.local"]

- service: kyuubi
  componentList:
    - component: service
      hostList: ["xinode2.local", "xinode3.local"]

- service: hbase
  componentList:
    - component: client
      # 所有机器都安装client
      hostList: ["*"]
    - component: master
      # master 需要部署在hdfs的client节点
      hostList: ["node1.local", "node2.local"]
    - component: regionserver
      # 需要部署在datanode的节点
      hostList: ["node{3-4}.local"]
  configList:
    env: 
      "HBASE_MASTER_HEAPSIZE": "2g"

- service: redis_sentinel
  componentList:
    - component: server
      hostList: ["xinode4.local"]
    - component: slave
      hostList: ["xinode2.local","xinode3.local"]
    - component: sentinel
      hostList: ["xinode2.local","xinode3.local","xinode4.local"]

- service: hadoop_meta
  componentList:
    - component: service
      hostList: ["xinode1.local","xinode2.local"]
    # 部署在YARN ResourceManager节点上
    - component: scheduler
      hostList: ["xinode3.local", "xinode4.local"]
    # 需要部署在kerberos的master节点, 且要安装hdfs-client，仅支持单点部署
    - component: kdc
      hostList: ["xinode2.local"]

# 当前仅支持单点部署
- service: meta_service
  componentList:
    - component: service
      hostList: ["xinode1.local"]
  configList:
    env:
      "JAVA_OPTS": "-Xmx512M -Xms512M -server"

- service: easyeagle
  componentList:
    # 仅支持单点安装 backend
  - component: backend
    hostList: ["xinode3.local"]
  - component: parseservice
    # 选择default_yarn实例下任意一个装有yarn client组件的节点，仅支持单点安装 parseservice;如果没有defalt_yarn，请咨询部署开发
    hostList: ["xinode3.local"]
  - component: collector
    # 选择default_yarn实例下所有装有yarn nodemanager组件的节点;如果没有defalt_yarn，请咨询部署开发
    hostList: ["xinode5.local","xinode6.local", "xinode7.local", "xinode8.local"]

- service: bdms_meta
  componentList:
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-Xms512M -Xmx512M -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"  

- service: mammut
  componentList:
    - component: executor
      hostList: ["xinode1.local", "xinode2.local"]
    - component: webserver
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    executor: 
      "JAVA_OPTS": "-Xmx1g -Xms1g -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"  
    webserver:
      "JAVA_OPTS": "-Xmx1g -Xms1g -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"  

- service: azkaban
  componentList:
    - component: exec
      hostList: ["xinode3.local", "xinode4.local"]
      volumeList:
        - <数据盘挂载路径 建议配置单独磁盘>
    - component: fc
      hostList: ["xinode3.local", "xinode4.local"]
    - component: web
      hostList: ["xinode3.local", "xinode4.local"]
    - component: lib
      hostList: ["xinode3.local", "xinode4.local"]
  configList:
    exec: 
      "AZKABAN_OPTS": "-Xmx3G -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"    
    web:
      "AZKABAN_OPTS": "-Xmx3G -Xms3G -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"

- service: easy_account
  version: 3.7.5.4
  componentList:
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_alert
  componentList:
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      JAVA_OPTS: "-server -Xmx1g -Xms1g -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"  

- service: easy_console
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_webmaster
  componentList:
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    nginx:
      # nginx 单点模式通过nginx ip 地址访问，nginx_ha默认是通过vip地址访问，可指定域名
      nginx_server_name: "http://<IP或域名>:11062"

- service: easy_aac
  componentList:
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_ddl
  componentList:
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_access
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: client
      # 部署在azkaban executor 节点
      hostList: ["xinode3.local", "xinode4.local"]
  configList:
    env:
      JAVA_OPTS: "-Xmx8G -Xms8G -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M -XX:ParallelGCThreads=35 -XX:+UseConcMarkSweepGC -XX:CMSInitiatingOccupancyFraction=70 -XX:+UseCMSInitiatingOccupancyOnly  -XX:+ParallelRefProcEnabled -XX:+PrintHeapAtGC -XX:NewSize=4g -XX:MaxNewSize=4g" 

- service: easy_metahub
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx4g -Xms4g -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"

- service: easy_transfer
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: client
      # 部署在azkaban executor 节点
      hostList: ["xinode3.local", "xinode4.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx1g -Xms1g -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -XX:+PrintTenuringDistribution -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"   

- service: easy_dmap
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_coop
  version: 1.2.2.5.1
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-Xms512m -Xmx512m -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"  

- service: easy_flow
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: engine
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_static
  componentList:
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
      
- service: easy_submit
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
      
- service: easy_udf
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_metaweb
  componentList:
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]

- service: easy_openapi
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-Xmx1g -Xms1g -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M" 

- service: easy_taskops
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx4g -Xms4g -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"

- service: easy_qa
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx512m -Xms512m -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"  

- service: easy_design
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx1g -Xms1g -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M" 

- service: easy_index
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-Xms1g -Xmx1g -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M" 

- service: easy_tag
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-Xms1g -Xmx1g -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"

- service: easy_dqc
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: client
      # 部署在azkaban executor 节点
      hostList: ["xinode3.local", "xinode4.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xms1g -Xmx1g -XX:MaxPermSize=128m -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"

- service: easy_test
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: client
      # 部署在azkaban executor 节点
      hostList: ["xinode3.local", "xinode4.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx512m -Xms512m -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -XX:+PrintTenuringDistribution -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M" 

- service: easy_standard
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]


# -----------------数据服务-----------------
- service: kong
  componentList:
    - component: cassandra
      hostList: ["xinode2.local", "xinode3.local"]
    - component: kong
      hostList: ["xinode3.local", "xinode4.local"]
    - component: konga
      # 仅支持单点部署
      hostList: ["xinode3.local"]

- service: easy_dataservice
  componentList:
    - component: backend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: frontend
      hostList: ["xinode1.local", "xinode2.local"]
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]
    - component: monitor
      # 此组件仅支持单点部署
      hostList: ["xinode1.local"]
    - component: orcserver
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      "JAVA_OPTS": "-server -Xmx1g -Xms1g -XX:SurvivorRatio=8 -Xss256k -XX:PermSize=128m -XX:MaxPermSize=128m -XX:+UseParNewGC -XX:MaxTenuringThreshold=15 -XX:+UseConcMarkSweepGC -XX:+CMSParallelRemarkEnabled -XX:+UseCMSCompactAtFullCollection -XX:+CMSClassUnloadingEnabled -XX:+UseCMSInitiatingOccupancyOnly -XX:CMSInitiatingOccupancyFraction=75 -XX:-DisableExplicitGC -XX:+UnlockDiagnosticVMOptions -XX:ParGCCardsPerStrideChunk=4096 -XX:-UseBiasedLocking -verbose:gc -XX:+PrintGCDateStamps -XX:+PrintGCTimeStamps -XX:+PrintGCDetails -XX:+HeapDumpOnOutOfMemoryError -XX:HeapDumpPath=$LOGS_DIR/ -Xloggc:$LOGS_DIR/gc-%t.log -XX:+UseGCLogFileRotation -XX:NumberOfGCLogFiles=2 -XX:GCLogFileSize=3M"
# -----------------数据服务-----------------


# -----------------数据资产-----------------
- service: meta_worker
  componentList:
    - component: api_server
      hostList: ["xinode1.local"]
    - component: meta_server
      hostList: ["xinode1.local"]
  configList:
    env:
      "META_WORKER_HEAPSIZE": "-Xmx4g -Xms4g -Xmn2g"  

- service: smilodon_fsimage_audit
  componentList:
    - component: fsimage_oiv
      hostList: ["xinode1.local", "xinode2.local"]
    - component: upload_audit
      # 部署在 HDFS NameNode 节点
      hostList: ["xinode3.local", "xinode4.local"]
    - component: upload_fsimage
      # 部署在 HDFS NameNode 节点
      hostList: ["xinode3.local", "xinode4.local"]
# -----------------数据资产-----------------


# -----------------实时组件-----------------
- service: ds_agent
  componentList:
    - component: agent
      # 部署在 YARN NodeManager 节点
      hostList: ["xinode5.local", "xinode6.local", "xinode7.local", "xinode8.local"]
  configList:
    ds_agent:
      "agent_jvm_conf": "-Xms256m -Xmx256m"  

- service: grafana
  componentList:
    - component: server
      hostList: ["xinode1.local"]

- service: ntsdb
  componentList:
    - component: master
      hostList: ["xinode2.local", "xinode3.local", "xinode4.local"]
    - component: shardserver
      hostList: ["xinode2.local", "xinode3.local", "xinode4.local"]

- service: logstash
  componentList:
    - component: server
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    env:
      Xms: "1g"
      Xmx: "1g"  

- service: realtime_debugger
  componentList:
    - component: plugin_server
      hostList: ["xinode1.local", "xinode2.local"]
  dependenceList:
    - ne-flink-1.10.0
    - ne-flink-1.12.4
    - ne-flink-1.13.3
    - ne-flink-1.14.0
    - plugin_cdc_ne-flink-1.13.3
    - plugin_ne-flink-1.10.0
    - plugin_ne-flink-1.12.4
    - plugin_ne-flink-1.14.0
    - ndi_ne-flink-1.13.3
  configList:
    plugin_server:
      "java_opts": "-Xms1g -Xmx1g" 

- service: realtime_monitor
  componentList:
    - component: monitor
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    monitor:
      "java_opts": "-Xms2g -Xmx2g" 

- service: realtime_ops
  componentList:
    - component: ops
      hostList: ["xinode1.local", "xinode2.local"]
    - component: web
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    ops:
      "java_opts": "-Xms1g -Xmx1g" 

- service: realtime_portal
  componentList:
    - component: portal
      hostList: ["xinode1.local", "xinode2.local"]
  configList:
    portal:
      "java_opts": "-Xms1g -Xmx1g"

- service: realtime_submitter
  componentList:
    - component: submitter
      hostList: ["xinode1.local", "xinode2.local"]
  dependenceList:
    - ne-flink-1.10.0
    - ne-flink-1.12.4
    - ne-flink-1.13.3
    - ne-flink-1.14.0
    - plugin_cdc_ne-flink-1.13.3
    - plugin_ne-flink-1.10.0
    - plugin_ne-flink-1.12.4
    - plugin_ne-flink-1.14.0
    - ndi_ne-flink-1.13.3
  configList:
    submitter:
      "java_opts": "-Xms2g -Xmx2g"

- id: ne-flink-1.10.0
  name: ne-flink-1.10.0
  service: flink
  version: ne-flink-1.10.0-1.1.8-scala-2.12
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: ne-flink-1.12.4
  name: ne-flink-1.12.4
  service: flink
  version: ne-flink-1.12.4-1.1.7.1-scala-2.12
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: ne-flink-1.13.3
  name: ne-flink-1.13.3
  service: flink
  version: ne-flink-1.13.3-1.0.1-scala-2.11
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: ne-flink-1.14.0
  name: ne-flink-1.14.0
  service: flink
  version: ne-flink-1.14.0-1.0.4-scala-2.12
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: plugin_cdc_ne-flink-1.13.3
  name: plugin_cdc_ne-flink-1.13.3
  service: flink_plugin
  version: cdc_ne-flink-1.13.3-1.0.1_scala2.11-release-3.9.4-2.1.7
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: plugin_ne-flink-1.10.0
  name: plugin_ne-flink-1.10.0
  service: flink_plugin
  version: ne-flink-1.10.0-1.1.8_scala2.12_hive2.1.1-release-3.9.4-1.4.5
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: plugin_ne-flink-1.12.4
  name: plugin_ne-flink-1.12.4
  service: flink_plugin
  version: ne-flink-1.12.4-1.1.7.1_scala2.12_hive2.1.1-release-3.9.4-1.4.5
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: plugin_ne-flink-1.14.0
  name: plugin_ne-flink-1.14.0
  service: flink_plugin
  version: ne-flink-1.14.0-1.0.4_scala2.12_hive2.1.1-release-3.9.4-1.4.5.2
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- id: ndi_ne-flink-1.13.3
  name: ndi_ne-flink-1.13.3
  service: flink_plugin
  version: ndi_ne-flink-1.13.3-1.0.1_scala2.11-release-3.0.0-1.0.0
  componentList:
    - component: client
      hostList: ["xinode1.local","xinode2.local"]

- service: sloth
  componentList:
    - component: server
      hostList: ["xinode1.local","xinode2.local"]
    - component: develop_web
      hostList: ["xinode1.local","xinode2.local"]
  dependenceList:
    - ne-flink-1.10.0
    - ne-flink-1.12.4
    - ne-flink-1.13.3
    - ne-flink-1.14.0     
    - plugin_cdc_ne-flink-1.13.3
    - plugin_ne-flink-1.10.0
    - plugin_ne-flink-1.12.4
    - plugin_ne-flink-1.14.0
    - ndi_ne-flink-1.13.3
  configList:
    server:
      "java_opts": "-Xms2g -Xmx2g" """

def first_sheet(wb):
    return wb[wb.sheetnames[0]]

def string_convert(str):
    return str.lower().strip(" ").replace("_", "")

# first arguments is xlsx file name
xlsx_file = sys.argv[1]
# get service_name and host_name
# print("I will convert the excel file to an yml file!")
# xlsx_file = input("Please Enter Excel filepath: ")
if not os.path.exists(xlsx_file):
    print(f"File {xlsx_file} does not exist")
    exit(1)
wb = openpyxl.load_workbook(xlsx_file)
ws = first_sheet(wb)
maxC = ws.max_column
maxR = ws.max_row
services = []
hostnames = []
for col_num in range(6, maxC + 1):
    service_name = ws.cell(row=3, column=col_num).value
    services.append(string_convert(service_name))
print(F"All Services in {xlsx_file}:")
print(services)
if 'sloth' in services and 'flink' not in services:
    services.append('flink')
    services.append('flinkplugin')
for row_num in range(4, maxR + 1):
    hostname = ws.cell(row=row_num, column=3).value
    hostnames.append(hostname.strip())
# print(hostnames)


# 删除yaml中不存在的service
yaml = ruamel.yaml.YAML()
with open('abcdefgdddd.yml', 'w', encoding='utf8') as f:
    f.write(strstr)
    f.flush()
with open('abcdefgdddd.yml', encoding='utf8') as f:
    yml_obj = yaml.load(f)
os.remove('abcdefgdddd.yml')
    
# with open('easyops7.0.yml', encoding='utf8') as f:
#     yml_obj = yaml.load(f)
new_yml_obj = []
for service in yml_obj:
    # print(string_convert(service.get('service')))
    sn = string_convert(service.get('service'))
    if sn in services:
        new_yml_obj.append(service)
    else:
        print(f"Removed service: {service.get('service')}")
# print all service name in the yml file
new_services = []
for service in new_yml_obj:
    service_name = service.get('service')
    new_services.append(service_name)
print()
print("All service in the xwphs.yml:")
print(new_services)
# component 对应主机数组
comp_list = []
servicess = ['kyuubi', 'metaservice']
server = ['neo4j', 'knox', 'zookeeper', 'bdmsmeta', 'easyaac', 'easyaccount', 'easyalert', 'easyddl', 'easydataservice', 'grafana', 'logstash', 'sloth']
admin = ['easyranger']
agent = ['dsagent']
engine = ['easyflow']
orcserver = ['easydataservice']
ops = ['realtimeops']
web = ['realtimeops']
develop_web = ['sloth']
first_comp = ['realtimedebugger', 'realtimemonitor', 'realtimesubmitter', 'realtimeportal']
frontend = ['easywebmaster', 'easyconsole', 'easyaccess', 'easydmap', 'easytransfer', 'easycoop', 'easymetaweb', 'easytaskops', 'easydqc', 'easydesign', 'easydataservice', 'easyindex', 'easystandard', 'easyflow', 'easystatic', 'easysubmit', 'easyudf', 'easyqa', 'easytag', 'easytest']
backend = ['easyconsole', 'easyaccess', 'easydmap', 'easytransfer', 'easymetahub', 'easycoop', 'easyopenapi', 'easytaskops', 'easydqc', 'easydesign', 'easydataservice', 'easyindex', 'easystandard', 'easyeagle', 'easyflow', 'easysubmit', 'easyudf', 'easyqa', 'easytag', 'easytest']
sloth_flink = []
nodem_coll = []
for col_num in range(6, maxC + 1):
    comp_dict = {}
    for row_num in range(4, maxR + 1):
        cell = ws.cell(row=row_num, column=col_num)
        if cell.value != None:
            components = cell.value.lower().strip(" ").split()
            # comp_dict = {}
            host_name = ws.cell(row=row_num, column=3).value.strip(" ")
            for component in components:
                comp_dict.setdefault(component, []).append(host_name)
    if 'sloth' in comp_dict:
        sloth_flink = comp_dict.get('sloth')
    if 'nodemanager' in comp_dict:
        nodem_coll = comp_dict.get('nodemanager')
    # print(comp_dict)
    for service in new_yml_obj:
        service_name = string_convert(service.get('service'))
        # print(service_name)
        if service_name == string_convert(ws.cell(row=3, column=col_num).value):
            for key in comp_dict.keys():
                for component in service.get('componentList'):
                    if key == component.get('component'):
                        component['hostList'] = comp_dict.get(key)
                    if key in servicess and component.get('component') == 'service':
                        component['hostList'] = comp_dict.get(key)
                    if key in server and component.get('component') == 'server':
                        component['hostList'] = comp_dict.get(key)
                    if key in admin and component.get('component') == 'admin':
                        component['hostList'] = comp_dict.get(key)
                    if key in frontend and component.get('component') == 'frontend':
                        component['hostList'] = comp_dict.get(key)
                    if key in backend and component.get('component') == 'backend':
                        component['hostList'] = comp_dict.get(key)
                    if key in orcserver and component.get('component') == 'orcserver':
                        component['hostList'] = comp_dict.get(key)
                    if key in agent and component.get('component') == 'agent':
                        component['hostList'] = comp_dict.get(key)
                    if key in engine and component.get('component') == 'engine':
                        component['hostList'] = comp_dict.get(key)
                    if key in ops and component.get('component') == 'ops':
                        component['hostList'] = comp_dict.get(key)
                    if key in web and component.get('component') == 'web':
                        component['hostList'] = comp_dict.get(key)
                    if key in develop_web and component.get('component') == 'develop_web':
                        component['hostList'] = comp_dict.get(key)
                    if component.get('component') == 'collector':
                        component['hostList'] = nodem_coll
                if key in first_comp:
                    service.get('componentList')[0]['hostList'] = comp_dict.get(key)
if 'sloth' in services:
    # 修改flink和flink_plugin
    for service in new_yml_obj:
        service_name = string_convert(service.get('service'))
        if service_name == "flink" or service_name == "flinkplugin":
            service.get('componentList')[0]['hostList'] = sloth_flink

with open('xwphs.yml', 'w', encoding='utf8') as f:
    yaml.dump(new_yml_obj, f)
print()
print("yml file is xwphs.yml in the current directory!")
# input("Enter to Exit")
